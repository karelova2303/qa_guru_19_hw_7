import csv
import os
import zipfile
from io import TextIOWrapper
from openpyxl import load_workbook

from pypdf import PdfReader

# Получение пути к архиву
archive_path = os.path.join(os.getcwd(), 'resources', 'archive_files.zip')


def test_csv(create_archive):
    with zipfile.ZipFile(archive_path) as zip_file:
        with zip_file.open('Test_CSV.csv') as csv_file:
            csvreader = list(csv.reader(TextIOWrapper(csv_file, 'utf-8-sig')))
            print(csvreader)

            assert csvreader == [['a1;b1;c1'],
                                 ['a2;b2;c2'],
                                 ['a3;b3;c3'],
                                 ['a4;b4;c4'],
                                 ['a5;b5;c5']]


def test_xlsx(create_archive):
    with zipfile.ZipFile(archive_path) as zip_file:
        with zip_file.open('Test_Excel.xlsx') as xlsx_file:
            xlsxreader = load_workbook(xlsx_file).active
            excel_value = [cell.value for row in xlsxreader.iter_rows() for cell in row]
            print(excel_value)

            assert excel_value == ['a1', 'b1', 'c1', 'a2', 'b2', 'c2', 'a3', 'b3', 'c3', 'a4', 'b4', 'c4', 'a5', 'b5',
                                   'c5']


def test_pdf(create_archive):
    with zipfile.ZipFile(archive_path) as zip_file:
        with zip_file.open('Test_Pdf.pdf') as pdf_file:
            pdfreader = PdfReader(pdf_file)
            text = pdfreader.pages[0].extract_text()
            print(text)

            assert 'This is a test PDF document.' in text
